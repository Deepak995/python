import openpyxl
import os
# import re
# import logging
import pdb
dict = {'GigabitEthernet0/0': {'ethernet_port': 'GigabitEthernet0/0', 'int_type': 'Physical', 'vlan': '-', 'mac_address': '00:d7:8f:a0:a8:00', 'poe': 'Disabled', 'speed': '100', 'mac_mtu': '1500', 'operational_status': 'reachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/1': {'ethernet_port': 'GigabitEthernet2/0/1', 'int_type': 'Physical', 'vlan': '-', 'mac_address': '00:d7:8f:a0:a8:42', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'reachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/10': {'ethernet_port': 'GigabitEthernet2/0/10', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0a', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/11': {'ethernet_port': 'GigabitEthernet2/0/11', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0b', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/12': {'ethernet_port': 'GigabitEthernet2/0/12', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0c', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/13': {'ethernet_port': 'GigabitEthernet2/0/13', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0d', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/14': {'ethernet_port': 'GigabitEthernet2/0/14', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0e', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/15': {'ethernet_port': 'GigabitEthernet2/0/15', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:0f', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/16': {'ethernet_port': 'GigabitEthernet2/0/16', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:10', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/2': {'ethernet_port': 'GigabitEthernet2/0/2', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:02', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/3': {'ethernet_port': 'GigabitEthernet2/0/3', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:03', 'poe': 'Disabled', 'speed': '100', 'mac_mtu': '1500', 'operational_status': 'reachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/4': {'ethernet_port': 'GigabitEthernet2/0/4', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:04', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/5': {'ethernet_port': 'GigabitEthernet2/0/5', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:05', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/6': {'ethernet_port': 'GigabitEthernet2/0/6', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:06', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/7': {'ethernet_port': 'GigabitEthernet2/0/7', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:07', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/8': {'ethernet_port': 'GigabitEthernet2/0/8', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:08', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'GigabitEthernet2/0/9': {'ethernet_port': 'GigabitEthernet2/0/9', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:09', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/17': {'ethernet_port': 'TenGigabitEthernet2/0/17', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:11', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/18': {'ethernet_port': 'TenGigabitEthernet2/0/18', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:12', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/19': {'ethernet_port': 'TenGigabitEthernet2/0/19', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:13', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/20': {'ethernet_port': 'TenGigabitEthernet2/0/20', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:14', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/21': {'ethernet_port': 'TenGigabitEthernet2/0/21', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:15', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/22': {'ethernet_port': 'TenGigabitEthernet2/0/22', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:16', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'reachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/23': {'ethernet_port': 'TenGigabitEthernet2/0/23', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:17', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/0/24': {'ethernet_port': 'TenGigabitEthernet2/0/24', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:18', 'poe': 'Disabled', 'speed': '1000', 'mac_mtu': '1500', 'operational_status': 'reachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/1/1': {'ethernet_port': 'TenGigabitEthernet2/1/1', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:19', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/1/2': {'ethernet_port': 'TenGigabitEthernet2/1/2', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:1a', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/1/3': {'ethernet_port': 'TenGigabitEthernet2/1/3', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:1b', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}, 'TenGigabitEthernet2/1/4': {'ethernet_port': 'TenGigabitEthernet2/1/4', 'int_type': 'Physical', 'vlan': 'default (1)', 'mac_address': '00:d7:8f:a0:a8:1c', 'poe': 'Disabled', 'speed': '10000', 'mac_mtu': '1500', 'operational_status': 'unreachable', 'admin_status': 'reachable'}}
# # st = 'WS-C3850-12X48U, WS-C3850-12X48U-E, WS-C3850-12X48U'
# # assert 'WS-C3850-12X48U' in  st, 'fail'
# # #print('yess')
console_stack =  {'1': {'serial': '1', 'role': 'Member', 'mac': '00:00:00:00:00:00', 'priority': '0'}, '2': {'serial': '2', 'role': 'Master', 'mac': '2c:86:d2:d9:92:00', 'priority': '9', 'Port 1': 'None', 'Port 2': 'None'}, '3': {'serial': '3', 'role': 'Member', 'mac': '00:00:00:00:00:00', 'priority': '0'}}
# ui_stack_info = {'2': {'member_no': '2', 'role': 'Master', 'mac_address': '2c:86:d2:d9:92:00', 'state': 'Ready', 'priority': '9', 'switch_ports': {'switch_port': '2/1', 'neighbor_port': ['NONE', 'NONE']}},
# '1': {'member_no': '1', 'role': 'Member', 'mac_address': '00:00:00:00:00:00', 'state': 'Provisioned', 'priority': '0', 'switch_ports': '-'},
# '3': {'member_no': '3', 'role': 'Member', 'mac_address': '00:00:00:00:00:00', 'state': 'Provisioned', 'priority': '0', 'switch_ports': '-'}}
# for key, value in ui_stack_info.items():
#     if key in console_stack:
#         logging.info("\t Stack state : UI Side: {}, Console Side: {}".format(ui_stack_info[key]['role'], console_stack[key]['role']))
#         logging.info("\t Stack serial no : UI Side: {}, Console Side: {}".format(key, key))
#         logging.info("\t Mac Address : UI Side: {}, Console Side: {}".format(ui_stack_info[key]['mac_address'], console_stack[key]['mac']))
#         logging.info("\t Stack priority : UI Side: {}, Console Side: {}".format(ui_stack_info[key]['priority'], console_stack[key]['priority']))
#         assert key in console_stack, 'stack serial no. does not match'
#         assert ui_stack_info[key]['role'] == console_stack[key]['role'], 'stack state does not match'
#         assert ui_stack_info[key]['mac_address'] == console_stack[key]['mac'], 'Mac address does not match'
#         assert ui_stack_info[key]['priority'] == console_stack[key]['priority'], 'switch priority does not match'
#
#         if ui_stack_info[key]['switch_ports'] != '-':
#             assert ui_stack_info[key]['switch_ports']['neighbor_port'][0].lower() == console_stack[key]['Port 1'].lower()\
#                  and ui_stack_info[key]['switch_ports']['neighbor_port'][1].lower() == console_stack[key]['Port 2'].lower(), 'switch neighbor port does not match'
#
directory = 'D:/'
file = 'check'
dir1 = os.path.join(directory, file)
os.mkdir(dir1)
wb = openpyxl.Workbook()
sheet1 = wb.create_sheet()
print(wb.sheetnames)
sheet1 = wb.create_sheet('jaishree')
active_sheet = wb['jaishree']
active_sheet['A1'] = 'helloworld'
for count, key in enumerate(dict.keys(), start=1):
    #pdb.set_trace()
    print(key)
    active_sheet['A'+str(count)] = key

wb.save(dir1+'/prabhu.xlsx')
load_sheet = openpyxl.load_workbook(dir1+'/prabhu.xlsx')
active_sheet = load_sheet['jaishree']
lenn = len(active_sheet['A'])
output_key_len = len(console_stack)
pdb.set_trace()
for key, i in zip(console_stack.keys(), range(lenn+1, lenn+output_key_len+1)):
    active_sheet['A'+str(i)] = key
    print(active_sheet['A' + str(i)].value)
    pdb.set_trace()

load_sheet.save(dir1+'/prabhu.xlsx')


# lenn = len(active_sheet['A'])
# output_key_len = len(console_stack)
# pdb.set_trace()
# for key, i in zip(console_stack.keys(), range(lenn+1, lenn+output_key_len+1)):
#     active_sheet['A'+str(i)] = key
# wb.save(dir1+'/prabhu.xlsx')